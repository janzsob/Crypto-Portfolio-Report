import json
import boto3
import pandas as pd
from openpyxl import load_workbook
from pycoingecko import CoinGeckoAPI
from datetime import date
from fpdf import FPDF
import os
import smtplib
from email.message import EmailMessage
import matplotlib.pyplot as plt


s3 = boto3.client("s3")
bucket = "crypto-portfolio-bucket"
key = "portfolio.xlsx"

def lambda_handler(event, context):
    """ Collect new prices and append them in the excel file """
    # Download excel file from s3 bucket
    local_file = "/tmp/downloaded.xlsx"
    s3.download_file(bucket, key, local_file)
    
    # Make API call via to Coingeco for prices
    cg = CoinGeckoAPI()
    data = cg.get_price(ids='ethereum, binancecoin, cardano, ripple, 1inch', vs_currencies='usd')
    
    # Load excel by openpyxl
    wb = load_workbook("/tmp/downloaded.xlsx")
    ws = wb.worksheets[0]

    today = date.today() 

    # Append current prices in a new row
    ws.append([
        today, 
        round(data['ethereum']['usd'], 0), 
        round(data['binancecoin']['usd'], 0), 
        round(data['cardano']['usd'], 2), 
        round(data['ripple']['usd'], 2), 
        round(data['1inch']['usd'], 2)
        ])
        
    for table in ws.tables.values():
        if table.name == "Árak":
            table.ref = f"A1:F{ws.max_row}"
    
    # Calculate portfolio value
    ws_amounts = wb.worksheets[1]

    eth_amount = ws_amounts["B2"].value
    bnb_amount = ws_amounts["B3"].value
    ada_amount = ws_amounts["B4"].value
    xrp_amount = ws_amounts["B5"].value
    inch_amount = ws_amounts["B6"].value

    portfolio_value = round(
        (eth_amount * data['ethereum']['usd']) + 
        (bnb_amount * data['binancecoin']['usd']) + 
        (ada_amount * data['cardano']['usd']) + 
        (xrp_amount * data['ripple']['usd']) + 
        (inch_amount * data['1inch']['usd']),
    )

    # Append portfolio value in excel
    ws_value = wb.worksheets[2]

    ws_value.append([today, portfolio_value])

    # save excel file in folder tmp
    wb.save("/tmp/downloaded.xlsx")
    
    # Load back the excel file to s4 bucket
    s3.upload_file("/tmp/downloaded.xlsx", bucket, key)

    """" Create a chart for PDF """
    df = pd.read_excel("/tmp/downloaded.xlsx", "Érték")

    # columns of dates and portfolio value
    dates_vs = df['Dátum']
    values = df['Portfolió Érték']

    # It defines the style of figure
    plt.style.use('seaborn-darkgrid')

    plt.plot_date(dates_vs, values, linestyle='solid')

    # It rotates and formats dates on x axis
    plt.gcf().autofmt_xdate()

    plt.ylabel("Portfólió Értéke (USD)")

    # Adjusts the padding
    plt.tight_layout()

    # it saves the figure in png
    plt.savefig("/tmp/chart.png")


    """ Generate PDF file """
    # create pdf
    class PDF(FPDF):
        def header(self):
            self.set_font('helvetica', '', 14)
            # count weeks
            week_count = date.today().strftime("%V")
            self.cell(0, 5, f"{week_count}. hét, Crypto riport", border=0, align='L')
            # date
            today = date.today().strftime("%Y/%m/%d")
            self.cell(0, 5, today, border=0, align='R')
            # Line break
            self.ln(15)

    pdf = PDF()
    # default format for FPDF is A4
    # dimensions of A4: widt=210mm height=297mm
    pdf.add_page()
    pdf.set_font("Times", size=14)
    line_height = pdf.font_size * 2

    # Page1: portfolio composition

    # portfolio value and invested amount
    pdf.set_margins(left=50, top=10, right=50)

    # current portfolio value
    df = pd.read_excel("/tmp/downloaded.xlsx", "Érték")
    current_value = f"${str(df.iloc[-1,1])}" # extract portfolio value from the last row.
    # invested amount
    df = pd.read_excel("/tmp/downloaded.xlsx", "Befektetés")
    invested_amount = round(df.at[0, 'Befektetett összeg (USD)'])

    # building table
    pdf.set_font('helvetica', 'B', 14)
    col_width = pdf.epw / 2
    pdf.cell(col_width, line_height, "Befektetett összeg", border=0, align="C")
    pdf.cell(col_width, line_height, "Portfólió értéke", border=0, ln=1, align="C")

    pdf.set_font("Times", style="B", size=14)
    pdf.set_fill_color(252, 3, 3)  # background color
    pdf.cell(col_width, line_height, f"${invested_amount}", border=1, align="C", fill=True)
    pdf.set_fill_color(37, 196, 49)  # background color
    pdf.cell(col_width, line_height, current_value, border=1, align="C", fill=True)

    pdf.ln(16) # break

    # table about portfolio composition
    # header
    col_table_names = ["Valuták", "Mennyiség (db)"]
    pdf.set_font('helvetica', 'B', 14)
    col_width = pdf.epw / 2
    pdf.cell(pdf.epw, line_height, "Portfólió összetétele", border=0, ln=1, align="C")
    for title in col_table_names:
        pdf.set_font('helvetica', 'B', 13)
        pdf.cell(col_width, line_height, title, border=1, align="C")

    pdf.ln(line_height) # break
    pdf.set_font(style="")  # disabling bold text

    # table body
    table_content = []
    df = pd.read_excel("/tmp/downloaded.xlsx", "Mennyiség")
    pdf.set_font("Times", style="B", size=14)
    for index, rows in df.iterrows():
        row_list = [rows[0], str(rows[1])]
        table_content.append(row_list)

    # create table body
    for row in table_content:
        for item in row:
            pdf.set_font("Times", size=13)
            pdf.cell(col_width, line_height, item, border=1, align="C")
        pdf.ln(line_height)

    pdf.ln(8) # break

    # Insert a chart about changes in portfolio value
    pdf.set_font('helvetica', 'B', 14)
    pdf.cell(0, line_height, "Portfólió értékének alakulása", ln=1, align="C")
    pdf.set_margins(left=10, top=10, right=10) # default marging dimensions
    pdf.ln()
    pdf.image("/tmp/chart.png", x=5, y=135, w=210-10)
    
    # Page2: list of prices
    # load prices
    df = pd.read_excel('/tmp/downloaded.xlsx', "Árak")

    TABLE_COL_NAMES = list(df.columns)
    TABLE_DATA = []
    for index, rows in df.iterrows():
        row_list = [pd.Timestamp(rows['Dátum']).strftime("%Y/%m/%d") , f"${str(rows.Ethereum)}", f"${str(rows.BNB)}", f"${str(rows.Cardano)}", f"${str(rows.XRP)}", f"${str(rows['1inch'])}"]
        TABLE_DATA.append(row_list)

    # title for table
    pdf.add_page()
    pdf.set_font('helvetica', 'B', 14)
    pdf.cell(0, line_height, "Árak heti bontásban", ln=1, align="C")
    pdf.set_margins(left=10, top=0, right=10)

    # header in table
    col_width = pdf.epw / 6  # distribute content evenly
    pdf.set_font(style="B", size=14)  # enabling bold text
    for col_name in TABLE_COL_NAMES:
        if col_name == "Dátum":
            pdf.cell(col_width, line_height, '', border=1, align="C")
        else:
            pdf.cell(col_width, line_height, col_name, border=1, align="C")
    pdf.ln(line_height)
    pdf.set_font(style="")  # disabling bold text

    # table data
    for row in TABLE_DATA:
        for datum in row:
            pdf.set_font("Times", size=14)
            pdf.cell(col_width, line_height, datum, border=1, align="C")
        pdf.ln(line_height)

    # save pdf in folder tmp
    pdf.output("/tmp/report.pdf")
    
    
    """ Send report via email """
    email_address = os.environ.get("EMAIL")
    email_password = os.environ.get("EMAIL_PASS")

    msg = EmailMessage()
    msg['Subject'] = os.environ.get('SUBJECT')
    msg['From'] = email_address
    msg['To'] = [os.environ.get('RECIPIENTS')]
    msg['Cc'] = [email_address]

    msg.set_content(os.environ.get('MESSAGE'))

    with open("/tmp/report.pdf", 'rb') as f:
        file_data = f.read()
        file_name = f.name

    msg.add_attachment(file_data, maintype="application", subtype="pdf", filename="report.pdf")

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(email_address, email_password)
        smtp.send_message(msg)