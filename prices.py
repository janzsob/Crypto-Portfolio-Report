from pycoingecko import CoinGeckoAPI
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table


def creating_file():
    # Using Coingecki API
    cg = CoinGeckoAPI()
    data = cg.get_price(ids='ethereum, binancecoin, cardano, ripple, 1inch', vs_currencies='usd')

    today = date.today() 

    # workbook
    wb = Workbook()
    # worksheet
    ws = wb.active
    ws.title = "Árak"

    # add column headings. NB. these must be strings
    ws.append(["Dátum", "Ethereum", "BNB", "Cardano", "XRP", "1inch"])

    # appending data from coingecko api
    data = [
        [
        today, data['ethereum']['usd'], 
        data['binancecoin']['usd'], 
        data['cardano']['usd'], 
        data['ripple']['usd'], 
        data['1inch']['usd']
        ],
    ]

    for row in data:
        ws.append(row)

    # Creating table
    table = Table(displayName="Crypto_prices", ref="A1:F2")

    ws.add_table(table)
    wb.save("portfolio.xlsx")


""" Collecting historical data (price) and append it in a file"""
def gather_historic_data(date_request):
    cg = CoinGeckoAPI()

    # collecting prices
    request_eth = cg.get_coin_history_by_id(id="ethereum", date=date_request, localization=False)
    eth = request_eth['market_data']['current_price']['usd']
    
    request_bnb = cg.get_coin_history_by_id(id="binancecoin", date=date_request, localization=False)
    bnb = request_bnb['market_data']['current_price']['usd']
    
    request_ada = cg.get_coin_history_by_id(id="cardano", date=date_request, localization=False)
    ada = request_ada['market_data']['current_price']['usd']
    
    request_xrp = cg.get_coin_history_by_id(id="ripple", date=date_request, localization=False)
    xrp = request_xrp['market_data']['current_price']['usd']
    
    request_1inch = cg.get_coin_history_by_id(id="1inch", date=date_request, localization=False)
    inch = request_1inch['market_data']['current_price']['usd']

    # loading excel file
    wb = load_workbook("portfolio.xlsx")
    ws = wb.worksheets[0]

    day = int(date_request[0:2])
    if date_request[3] == '0':
        month = int(date_request[4])
    else:
        month = int(date_request[3:5])
    year = int(date_request[6:11])
    date_append = date(year, month, day) 

    ws.append([date_append, round(eth, 0), round(bnb, 0), round(ada, 2), round(xrp, 2), round(inch, 2)])

    for table in ws.tables.values():
        if table.name == "Árak":
            table.ref = f"A1:F{ws.max_row}"
    
    wb.save("portfolio.xlsx")

def get_historic_value(date_request):
    cg = CoinGeckoAPI()

    # collecting prices
    request_eth = cg.get_coin_history_by_id(id="ethereum", date=date_request, localization=False)
    eth = request_eth['market_data']['current_price']['usd']
    
    request_bnb = cg.get_coin_history_by_id(id="binancecoin", date=date_request, localization=False)
    bnb = request_bnb['market_data']['current_price']['usd']
    
    request_ada = cg.get_coin_history_by_id(id="cardano", date=date_request, localization=False)
    ada = request_ada['market_data']['current_price']['usd']
    
    request_xrp = cg.get_coin_history_by_id(id="ripple", date=date_request, localization=False)
    xrp = request_xrp['market_data']['current_price']['usd']
    
    request_1inch = cg.get_coin_history_by_id(id="1inch", date=date_request, localization=False)
    inch = request_1inch['market_data']['current_price']['usd']

    # Calculate portfolio value
    wb = load_workbook("portfolio.xlsx")
    ws_amounts = wb.worksheets[1]

    eth_amount = ws_amounts["B2"].value
    bnb_amount = ws_amounts["B3"].value
    ada_amount = ws_amounts["B4"].value
    xrp_amount = ws_amounts["B5"].value
    inch_amount = ws_amounts["B6"].value

    portfolio_value = round(
        (eth_amount * eth) + 
        (bnb_amount * bnb) + 
        (ada_amount * ada) + 
        (xrp_amount * xrp) + 
        (inch_amount * inch),
    )

    day = int(date_request[0:2])
    if date_request[3] == '0':
        month = int(date_request[4])
    else:
        month = int(date_request[3:5])
    year = int(date_request[6:11])
    date_append = date(year, month, day) 

    # Append portfolio value in excel
    ws_value = wb.worksheets[2]

    ws_value.append([date_append, portfolio_value])

    wb.save("portfolio.xlsx")

""" Loading workbook then appending data in new row """
def append_data():
    # Collect the current prices by CoinGecko API
    cg = CoinGeckoAPI()
    data = cg.get_price(ids='ethereum, binancecoin, cardano, ripple, 1inch', vs_currencies='usd')

    today = date.today()

    # append new prices
    wb = load_workbook("portfolio.xlsx")
    ws_prices = wb.worksheets[0]

    ws_prices.append([
        today, 
        round(data['ethereum']['usd'], 0), 
        round(data['binancecoin']['usd'], 0), 
        round(data['cardano']['usd'], 2), 
        round(data['ripple']['usd'], 2), 
        round(data['1inch']['usd'], 2)
        ])

    for table in ws_prices.tables.values():
        if table.name == "Árak":
            table.ref = f"A1:F{ws_prices.max_row}"
        #print(table.name)
        #print(table.ref)

    # Calculate portfolio value
    ws_amounts = wb.worksheets[1]

    eth_amount = ws_amounts["B2"].value
    bnb_amount = ws_amounts["B3"].value
    ada_amount = ws_amounts["B4"].value
    xrp_amount = ws_amounts["B5"].value
    inch_amount = ws_amounts["B6"].value

    portfolio_value = round(
        (eth_amount * data['ethereum']['usd']) + 
        (bnb_amount * data['binancecoin']['usd']) + 
        (ada_amount * data['cardano']['usd']) + 
        (xrp_amount * data['ripple']['usd']) + 
        (inch_amount * data['1inch']['usd']),
    )

    # Append portfolio value in excel
    ws_value = wb.worksheets[2]

    ws_value.append([today, portfolio_value])

    wb.save("portfolio.xlsx")


#gather_historic_data('05-09-2021')
#get_historic_value('05-09-2021')

#append_data()

